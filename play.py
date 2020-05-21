from openpyxl import Workbook, load_workbook
from ast import literal_eval

def parse_game(excel_name):
    game_wb = load_workbook(excel_name)
    game_ws = game_wb.active

    board = [[] for _ in range(24)]

    indexes_up = {
        2: 23,
        3: 22,
        4: 21,
        5: 20,
        6: 19,
        7: 18,
        10: 17,
        11: 16,
        12: 15,
        13: 14,
        14: 13,
        15: 12
    }

    indexes_down = {
        2: 0,
        3: 1,
        4: 2,
        5: 3,
        6: 4,
        7: 5,
        10: 6,
        11: 7,
        12: 8,
        13: 9,
        14: 10,
        15: 11
    }

    on_bar = {
        'white': [],
        'black': []
    }

    game_state = []

    hit_none = False
    for i in range(1, 17):
        for j in range(2, 20):
            if i in (1, 8, 9, 16):
                cell_val = game_ws.cell(row=j, column=i).value
                if cell_val == 1:
                    on_bar['white'].append('white')
                if cell_val == 0:
                    on_bar['black'].append('black')
            else:
                cell_val = game_ws.cell(row=j, column=i).value
                if cell_val == None:
                    hit_none = True
                if cell_val == 1:
                    if not hit_none:
                        board[indexes_up[i]].append('white')
                    else:
                        board[indexes_down[i]].append('white')
                if cell_val == 0:
                    if not hit_none:
                        board[indexes_up[i]].append('black')
                    else:
                        board[indexes_down[i]].append('black')
        hit_none = False

    white_count = 0
    black_count = 0
    for cell in board:
        for checker in cell:
            if checker == 'white':
                white_count += 1
            else:
                black_count += 1

    off_board = {
        'white': ['white' for i in range(15 - white_count)],
        'black': ['black' for i in range(15 - black_count)]
    }

    pieces_left = {
        'white': white_count,
        'black': black_count
    }


    turn = 'white' if int(game_ws['H23'].value) == 1 else 'black' 
    roll = literal_eval(game_ws['I23'].value)

    game_state.append(board)
    game_state.append(on_bar)
    game_state.append(off_board)
    game_state.append(pieces_left)
    game_state.append(turn)
    game_state.append(roll)

    return game_state


def parse_weights(net_output_file_name):

    net_output_wb = load_workbook(net_output_file_name)
    weights_ws = net_output_wb.active

    input_range = weights_ws['A1':'AN198']
    hidden_range = weights_ws['A200':'B239']

    for 
