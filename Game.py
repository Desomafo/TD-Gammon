import numpy as np
from openpyxl import Workbook, load_workbook
from ast import literal_eval
from multiprocessing import Pool
import random

class Game:

    def __init__(self, first_opponent, second_opponent=None, is_learn=False):
        self.players = ['white', 'black']
        self.first_opponent = first_opponent
        if second_opponent == None:
            self.second_opponent = self.first_opponent
        else:
            self.second_opponent = second_opponent
        self.is_learn = is_learn
        self.roll = ()
        self.board = [[] for _ in range(24)]
        self.on_bar = {}
        self.off_board = {}
        self.pieces_left = {}
        self.turn = ''
        for p in self.players:
            self.on_bar[p] = []
            self.off_board[p] = []
            self.pieces_left[p] = 15

        # Initialize the board with a hard coded set-up for each point (0 indexed!)
        for i in range(2):
            self.board[0].append('white')
        for i in range(5):
            self.board[5].append('black')
        for i in range(3):
            self.board[7].append('black')
        for i in range(5):
            self.board[11].append('white')
        for i in range(5):
            self.board[12].append('black')
        for i in range(3):
            self.board[16].append('white')
        for i in range(5):
            self.board[18].append('white')
        for i in range(2):
            self.board[23].append('black')

    def roll_dice(self):
        return (random.randint(1,6), random.randint(1,6))


    def find_moves(self, roll, player):
        moves = []
        # Probably can get rid of this and just reference the indices of the roll
        r1, r2 = roll
        for r in roll:
            # Can we bear off?
            if player == 'white':
                num_white_pieces = 0
                # Check to see if all the stones are in the home base for white
                for i in range(0,24):
                    if len(self.board[i]) >= 1 and i >= 18:
                        if (self.board[i][0] == player):
                            num_white_pieces = num_white_pieces + len(self.board[i])

                        if (num_white_pieces == 15) or (len(self.off_board[player]) >=1):
                            for i in range(0,24):
                                if (len(self.board[i]) > 0) and (self.board[i][0] == player):
                                    if i >= 17 and (i + r) == 24:
                                        move = (i, 'off')
                                        moves.append(move)

            if player == 'black':
                num_black_pieces = 0
                # Check to see if all the stones are in the home base for black
                for i in range(0, 24):
                    if len(self.board[i]) >= 1 and i <= 5:
                        if (self.board[i][0] == player):
                            num_black_pieces = num_black_pieces + len(self.board[i])

                        if (num_black_pieces == 15) or (len(self.off_board[player]) >= 1):
                            for i in range(0, 24):
                                if(len(self.board[i]) > 0) and (self.board[i][0] == player):
                                    if i <= 5 and (i - r) == -1:
                                        move = (i, 'off')
                                        moves.append(move)

            # Are there pieces on the bar for the given player?
            if len(self.on_bar[player]) >= 1:
                # print("Player {} has pieces on the bar".format(self.on_bar[player]))

                # Does the point where we would put the piece have no pieces or is it controlled by the player?
                if player == 'white':
                    for i in range(0,6):
                        if r == i and len(self.board[i]) <= 1:
                            move = ('bar', r)
                            moves.append(move)
                        elif r == i and self.board[i][0] == player:
                            move = ('bar', r)
                            moves.append(move)

                if player == 'black':
                    for i in range(18,24):
                        if (24 - i) == r and len(self.board[24 - r]) <= 1:
                            move = ('bar', 24 - r)
                            moves.append(move)
                        elif (24 - i) == r and self.board[24 - r][0] == player:
                            move = ('bar', 24 - r)
                            moves.append(move)

            # Can the player make a valid move with the roll?
            for i in range(0,24):
                if len(self.board[i]) > 0:
                    if self.board[i][0] == player:
                        if player == 'white':
                            if i + r < 24:
                                if (len(self.board[i + r]) <= 1) or (self.board[i + r][0] == player):
                                    move = (i, i + r)
                                    moves.append(move)
                        if player == 'black':
                            if i - r >= 0:
                                if (len(self.board[i - r]) <= 1) or (self.board[i - r][0] == player):
                                    move = (i, i - r)
                                    moves.append(move)

        ##################################################################
        # Can the player make a valid move with the combination of rolls?
        ##################################################################
        combo = r1 + r2

        if player == 'white':
            num_white_pieces = 0
            # Check to see if all the stones are in the home base for white
            for i in range(0,24):
                if len(self.board[i]) >= 1 and i >= 18:
                    if (self.board[i][0] == player):
                        num_white_pieces = num_white_pieces + len(self.board[i])

                if (num_white_pieces == 15) or (len(self.off_board[player]) >=1):
                    for i in range(0,24):
                        if (len(self.board[i]) > 0) and (self.board[i][0] == player):
                            if i >= 17 and (i + combo) == 24:
                                move = (i, 'off')
                                moves.append(move)

        if player == 'black':
            num_black_pieces = 0
            # Check to see if all the stones are in the home base for black
            for i in range(0, 24):
                if len(self.board[i]) >= 1 and i <= 5:
                    if (self.board[i][0] == player):
                        num_black_pieces = num_black_pieces + len(self.board[i])

                if (num_black_pieces == 15) or (len(self.off_board[player]) >= 1):
                    for i in range(0, 24):
                        if(len(self.board[i]) > 0) and (self.board[i][0] == player):
                            if i <= 5 and (i - combo) == -1:
                                move = (i, 'off')
                                moves.append(move)



        # Are there pieces on the bar for the given player?
        if len(self.on_bar[player]) >= 1:
            # print("Player {} has pieces on the bar".format(self.on_bar[player]))

            # Does the point where we would put the piece have no pieces or is it controlled by the player?
            if player == 'white':
                for i in range(0,6):
                    if r == i and len(self.board[i]) <= 1:
                        move = ('bar', r)
                        moves.append(move)
                    elif r == i and self.board[i][0] == player:
                        move = ('bar', r)
                        moves.append(move)

            if player == 'black':
                for i in range(18,24):
                    if (24 - i) == r and len(self.board[24 - r]) <= 1:
                        move = ('bar', 24 - r)
                        moves.append(move)
                    elif (24 - i) == r and self.board[24 - r][0] == player:
                        move = ('bar', 24 - r)
                        moves.append(move)

        # Can the player make a valid move with the roll?
        for i in range(0,24):
            if len(self.board[i]) > 0:
                if self.board[i][0] == player:
                    if player == 'white':
                        if i + combo < 24:
                            if (len(self.board[i + combo]) <= 1) or (self.board[i + combo][0] == player):
                                move = (i, i + combo)
                                moves.append(move)
                    if player == 'black':
                        if i - combo >= 0:
                            if (len(self.board[i - combo]) <= 1) or (self.board[i - combo][0] == player):
                                move = (i, i - combo)
                                moves.append(move)

        # print("roll was {}".format(roll))
        return(moves)

    def take_action(self, player, move):
        # If there is a piece on the bar then move that first
        start, end = move
        if end == 'off':
            moved_piece = self.board[start].pop()
            self.off_board[player].append(moved_piece)
            self.pieces_left[player] = self.pieces_left[player] - 1

        elif start == 'bar':
            if len(self.board[end]) == 1:
                # Check for hit
                if self.board[end][0] != player:
                    hit_piece = self.board[end].pop()
                    self.on_bar[self.get_opponent(player)].append(hit_piece)
                    moved_piece = self.on_bar[player].pop()
                    self.board[end].append(moved_piece)
                else:
                    moved_piece = self.on_bar[player].pop()
                    self.board[end].append(moved_piece)

            elif len(self.board[end]) > 1 and self.board[end][0] == player:
                moved_piece = self.on_bar[player].pop()
                self.board[end].append(moved_piece)
            elif len(self.board[end]) == 0:
                moved_piece = self.on_bar[player].pop()
                self.board[end].append(moved_piece)
        else:
            # Normal Moves
            if len(self.board[end]) >= 1 and self.board[end][0] == player:
                moved_piece = self.board[start].pop()
                self.board[end].append(moved_piece)
            if len(self.board[end]) == 1 and self.board[0] != player:
                # Check for hit
                hit_piece = self.board[end].pop()
                self.on_bar[self.get_opponent(player)].append(hit_piece)
                moved_piece = self.board[start].pop()
                self.board[end].append(moved_piece)
            if len(self.board[end]) == 0:
                moved_piece = self.board[start].pop()
                self.board[end].append(moved_piece)

    def undo_action(self, player, move):
        start, end = move

        if end == 'off':
            moved_piece = self.off_board[player].pop()
            self.board[start].append(moved_piece)
            self.pieces_left[player] += 1

        elif start == 'bar':
            moved_piece = self.board[end].pop()
            self.on_bar[player].append(moved_piece)
        else:
            moved_piece = self.board[end].pop()
            self.board[start].append(moved_piece)

    # Check to see if the game is over
    def game_over(self):
        for p in self.players:
            if len(self.off_board[p]) == 15 and self.pieces_left[p] == 0:
                return True

    # Get the opposite color of the given player (useful for hitting)
    def get_opponent(self, player):
        for p in self.players:
            if p != player:
                return p

    # Determine which player has won
    def find_winner(self):
        if (len(self.off_board[self.players[0]]) == 15 and self.pieces_left[self.players[0]] == 0):
            return self.players[0]

        return self.players[1]

    # Print the contents of a desired point on the board
    def print_point(self, point):
        print("Point #{} has {} pieces: {}".format(point, len(self.board[point]), self.board[point]))


    def get_representation(self, board, players, on_bar, off_board, turn):
        # "There were a total of 198 input units..."
        network_input = []
        units = []
        # "For each point on the board, for units indicated the number of white pieces on the point"
        for p in players:
            for i in range(0, 24):
                if len(board[i]) >= 1:
                    if board[i][0] == p:
                        if len(board[i]) == 1:
                            units.append(1)
                            for i in range(1,4):
                                units.append(0)
                        elif len(board[i]) == 2:
                            for i in range(0,2):
                                units.append(1)
                            for i in range(2, 4):
                                units.append(0)
                        elif len(board[i]) == 3:
                            for i in range(0,3):
                                units.append(1)
                            units.append(0)
                        elif len(board[i]) > 3:
                            num_pieces = len(board[i])
                            for i in range(0, 3):
                                units.append(1)
                            units.append(float(((num_pieces - 3) / 2)))

                    else:
                        for i in range(0,4):
                            units.append(0)
                else:
                        for i in range(0,4):
                            units.append(0)


        units.append(len(on_bar['white']) / 2)
        units.append(len(on_bar['black']) / 2)

        if (len(off_board['white']) > 0):
            units.append(len(off_board['white']) / 15)
        else: units.append(0)

        if (len(off_board['black']) > 0):
            units.append(len(off_board['black']) / 15)
        else: units.append(0)

        if turn == 'white':
            units.append(1)
            units.append(0)
        else:
            units.append(0)
            units.append(1)

        return units


    def set_game_state(self, new_state):
        self.board = new_state[0]
        self.on_bar = new_state[1]
        self.off_board = new_state[2]
        self.pieces_left = new_state[3]
        self.turn = new_state[4]

    
    def find_best_action(self, actions):
        values = []
        repr_list = []

        # Find the action with the most appealing value
        for action in actions:
            self.take_action(self.turn, action)
            representation = self.get_representation(
                    self.board, self.players, self.on_bar,
                    self.off_board, self.turn
                    )
            repr_list.append(representation)
            # Undo the action and try the rest
            self.undo_action(self.turn, action)

        with Pool() as pool:
            try:
                if self.turn == 'white':
                    values = pool.map(self.first_opponent.getValue, repr_list)
                elif self.turn == 'black':
                    values = pool.map(self.second_opponent.getValue, repr_list)
            except KeyboardInterrupt:
                pool.terminate()
                print("Game is terminated")
                return

        # We want white to win so find the max for white and the smallest for black
        max_index = 0
        min_val = 1
        min_index = 0
        max_val = 0
        for i, (white_val, black_val) in enumerate(values):
            if self.turn == 'white':
                if max_val < white_val:
                    max_val = white_val
                    max_index = i
            elif self.turn == 'black':
                if max_val < black_val:
                    max_val = black_val
                    max_index = i
        best_action = actions[max_index]

        if self.is_learn == True:
            for i in range(0, len(values)):
                if self.turn == 'black':
                    if min_val > values[i][1]:
                        min_val = values[i][1]
                        min_index = i
            best_action = actions[min_index]

        return best_action



    def play(self):
        
        # print("White player rolled {}, Black player rolled {}".format(p1Roll[0] + p1Roll[1], p2Roll[0] + p2Roll[1]))
        p1Roll = (0,0)
        p2Roll = (0,0)
        while sum(p1Roll) == sum(p2Roll):
            p1Roll = self.roll_dice()
            p2Roll = self.roll_dice()

        if sum(p1Roll) > sum(p2Roll):
            print("White player gets the first turn...")
            self.turn = self.players[0]
        else:
            print("Black player gets the first turn")
            self.turn = self.players[1]
        start = 1
        moves = 0
        states = []

        while not self.game_over():
            actions = []

            if start == 1:
                actions = self.find_moves(p1Roll, self.turn)
                start = 0
            else:
                actions = self.find_moves(self.roll_dice(), self.turn)

            if len(actions) > 0:

                best_action = self.find_best_action(actions)
                # Take the best action
                self.take_action(self.turn, best_action)

                # Get the representation
                expected_board = self.get_representation(self.board, self.players, self.on_bar, self.off_board, self.turn)
                if self.turn == 'white':
                    # Save the state
                    states.append(expected_board)
                    # print(new_ANN.getValue(expected_board))
                    # print('state size',len(states))
                # Swap turns and increment move count
                moves += 1
                self.turn = self.get_opponent(self.turn)
                reward = 0
                if self.game_over():
                    print("Game over in {} moves".format(moves))
                    print("Num states: ", len(states))
                    print("{} won".format(self.find_winner()))

                    if self.find_winner() == 'white':
                        reward = 1
                    for i in range(len(self.board)):
                        self.print_point(i)

        return self.find_winner(), states


    def parse_game(self, xlsx_name):
        game_wb = load_workbook(xlsx_name)
        game_ws = game_wb.active

        board = [[] for _ in range(24)]

        indexes_up = {
            2: 23,
            3: 22,
            4: 21,
            5: 20,
            6: 19,
            7: 18,
            10: 17,
            11: 16,
            12: 15,
            13: 14,
            14: 13,
            15: 12
        }

        indexes_down = {
            2: 0,
            3: 1,
            4: 2,
            5: 3,
            6: 4,
            7: 5,
            10: 6,
            11: 7,
            12: 8,
            13: 9,
            14: 10,
            15: 11
        }

        on_bar = {
            'white': [],
            'black': []
        }

        game_state = []

        hit_none = False
        for i in range(1, 17):
            for j in range(2, 20):
                if i in (1, 8, 9, 16):
                    cell_val = game_ws.cell(row=j, column=i).value
                    if cell_val == 1:
                        on_bar['white'].append('white')
                    if cell_val == 0:
                        on_bar['black'].append('black')
                else:
                    cell_val = game_ws.cell(row=j, column=i).value
                    if cell_val == None:
                        hit_none = True
                    if cell_val == 1:
                        if not hit_none:
                            board[indexes_up[i]].append('white')
                        else:
                            board[indexes_down[i]].append('white')
                    if cell_val == 0:
                        if not hit_none:
                            board[indexes_up[i]].append('black')
                        else:
                            board[indexes_down[i]].append('black')
            hit_none = False

        white_count = 0
        black_count = 0
        for cell in board:
            for checker in cell:
                if checker == 'white':
                    white_count += 1
                else:
                    black_count += 1

        off_board = {
            'white': ['white' for i in range(15 - white_count)],
            'black': ['black' for i in range(15 - black_count)]
        }

        pieces_left = {
            'white': white_count,
            'black': black_count
        }


        turn = 'white' if int(game_ws['H23'].value) == 1 else 'black' 
        self.roll = literal_eval(game_ws['I23'].value)
        

        self.board = board
        self.on_bar = on_bar
        self.off_board = off_board
        self.pieces_left = pieces_left
        self.turn = turn
